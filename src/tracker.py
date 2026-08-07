"""Application tracking spreadsheet: one row per company+role, auto-updated on every run.

The Status column (application status: Applied / In Progress / Rejected / etc.) is
user-owned — this module never writes to it after the row's first creation.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config

TRACKER_PATH = config.DOCUMENTS_EXPORT_DIR / "Application Tracker.xlsx"

HEADERS = [
    "Date",
    "Company",
    "Role",
    "Platform",
    "Application Status",
    "Pipeline Result",
    "Score",
    "Pages",
    "Hiring Manager",
    "Notes",
    "Output Folder",
]
STATUS_OPTIONS = [
    "Not Applied",
    "Applied",
    "In Progress",
    "Interviewing",
    "Offer",
    "Rejected",
    "Withdrawn",
]
COL_WIDTHS = [12, 24, 34, 14, 16, 16, 8, 7, 14, 40, 50]


def _new_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_OPTIONS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add("E2:E1000")
    return wb


def _find_row(ws, company: str, role: str) -> int | None:
    for row in range(2, ws.max_row + 1):
        if (ws.cell(row=row, column=2).value or "") == company and (
            ws.cell(row=row, column=3).value or ""
        ) == role:
            return row
    return None


def upsert(
    *,
    company: str,
    role: str,
    platform: str,
    pipeline_result: str,
    score: int | None,
    pages: int | None,
    hiring_manager: str,
    notes: str,
    output_folder: str,
    run_date: str | None = None,
) -> None:
    """Add or refresh a row for this company+role. Never touches an existing Application Status."""
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(TRACKER_PATH) if TRACKER_PATH.exists() else _new_workbook()
    ws = wb["Applications"]

    row = _find_row(ws, company, role)
    if row is None:
        row = ws.max_row + 1
        ws.cell(row=row, column=5, value="Not Applied")

    values = [
        run_date or date.today().isoformat(),
        company,
        role,
        platform,
        None,  # Application Status — left alone on update, set above only when new
        pipeline_result,
        score,
        pages,
        hiring_manager,
        notes,
        output_folder,
    ]
    for col_idx, value in enumerate(values, start=1):
        if col_idx == 5:  # never overwrite Application Status
            continue
        ws.cell(row=row, column=col_idx, value=value)

    wb.save(TRACKER_PATH)
